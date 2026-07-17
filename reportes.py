"""
reportes.py
============
Generacion de reportes de asistencia en Excel (.xlsx) y PDF, listos para
imprimir o descargar. Ambas funciones reciben la misma lista de filas
(dicts) que ya se muestra en la tabla de "Registros" y devuelven los
bytes del archivo generado en memoria (no se guarda nada en disco).
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

NOMBRE_SISTEMA = "Island"

COLUMNAS_REPORTE = [
    ("fecha", "Fecha"),
    ("hora", "Hora"),
    ("id_estudiante", "Código"),
    ("nombre_completo", "Estudiante"),
    ("codigo_seccion", "Sección"),
    ("tipo_evento", "Movimiento"),
    ("turno", "Turno"),
    ("estado_alerta", "Alerta"),
    ("detalle_alerta", "Detalle"),
]


def _valor(fila: dict, clave: str) -> str:
    valor = fila.get(clave, "")
    return "" if valor is None else str(valor)


def generar_excel(filas: list[dict], titulo: str = "Reporte de asistencia") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia"

    encabezado_fill = PatternFill(start_color="00A8E8", end_color="00A8E8", fill_type="solid")
    encabezado_font = Font(bold=True, color="FFFFFF")

    ws.merge_cells("A1:I1")
    ws["A1"] = f"{NOMBRE_SISTEMA} — {titulo}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  •  Total de registros: {len(filas)}"
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    fila_encabezado = 4
    for col_idx, (_, etiqueta) in enumerate(COLUMNAS_REPORTE, start=1):
        celda = ws.cell(row=fila_encabezado, column=col_idx, value=etiqueta)
        celda.font = encabezado_font
        celda.fill = encabezado_fill
        celda.alignment = Alignment(horizontal="center")

    for fila_idx, fila in enumerate(filas, start=fila_encabezado + 1):
        for col_idx, (clave, _) in enumerate(COLUMNAS_REPORTE, start=1):
            ws.cell(row=fila_idx, column=col_idx, value=_valor(fila, clave))

    ws.auto_filter.ref = f"A{fila_encabezado}:I{fila_encabezado + max(len(filas), 1)}"
    ws.freeze_panes = f"A{fila_encabezado + 1}"

    anchos = [12, 10, 12, 26, 10, 12, 12, 14, 30]
    for idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generar_pdf(filas: list[dict], titulo: str = "Reporte de asistencia") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
        leftMargin=1.2 * cm,
        rightMargin=1.2 * cm,
        title=f"{NOMBRE_SISTEMA} - {titulo}",
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "TituloIsland", parent=estilos["Title"], textColor=colors.HexColor("#00A8E8"), fontSize=18,
    )
    estilo_sub = ParagraphStyle("SubIsland", parent=estilos["Normal"], textColor=colors.HexColor("#666666"), fontSize=9)

    elementos = [
        Paragraph(f"{NOMBRE_SISTEMA} — {titulo}", estilo_titulo),
        Paragraph(
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}  •  Total de registros: {len(filas)}",
            estilo_sub,
        ),
        Spacer(1, 0.5 * cm),
    ]

    encabezados = [etiqueta for _, etiqueta in COLUMNAS_REPORTE]
    datos = [encabezados]
    for fila in filas:
        datos.append([_valor(fila, clave) for clave, _ in COLUMNAS_REPORTE])

    if len(datos) == 1:
        datos.append(["Sin registros" if i == 0 else "" for i in range(len(encabezados))])

    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#00A8E8")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F7FA")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ]
        )
    )
    elementos.append(tabla)
    doc.build(elementos)
    return buffer.getvalue()
